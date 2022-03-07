
from tkinter import *
from tkinter import messagebox
import tkinter.font as tkFont
import openpyxl
from os.path import exists

excel_path = "data.xlsx"


excelObj = openpyxl.load_workbook(excel_path)
sheet = excelObj.active


def getValues():
    rowsVal = [str(sheet[x][0].value) for x in range(2, sheet.max_row+1)]
    return rowsVal


def saveData():
    aadharNewValue = aadharValue.get()
    # print(sheet.max_row)
    values = getValues()
    if not aadharNewValue.isdigit():
        showMessage("Please enter correct aadhar number")
        quit()
    if len(aadharNewValue) != 12:
        showMessage("Please enter 12 digit aadhar number")
        enterNumber.delete(0, END)
        return
    if aadharNewValue in values:
        showMessage("Aadhar number Exists")
        enterNumber.delete(0, END)

    else:
        maxRows = sheet.max_row+1
        print(maxRows)
        sheet[f'A{maxRows}'] = aadharNewValue
        print(getValues())
        excelObj.save("data.xlsx")
        aadharListView.delete(0, 'end')
        values = getValues()
        for x in range(len(values)):
            valueExcel = values[x]
            if valueExcel != "None":
                aadharListView.insert(x, valueExcel)
        heading.pack(padx=50, pady=10)
        aadharNo.pack(padx=10, )
        aadharListView.pack(padx=20, pady=10)
        enterNumber.delete(0, END)


def nextButtonFun():
    newWindow = Toplevel(window)
    newWindow.title("New Window")
    newWindow.config(background="black")
    newWindow.geometry('500x500')
    Label(
        newWindow,
        text="Aadhar Verification",
        foreground="white",
        background="black",
        width=30,
        font=heading_stlye
    ).pack(padx=50, pady=10)
    Label(
        newWindow,
        text="Please Enter Aadhar No",
        foreground="white",
        background="black",
        width=30,
        height=1,
        font=subheading_stlye
    ).pack(padx=100, pady=20)
    Entry(
        newWindow, width=20,
        textvariable=NewaadharValue,
        font=('Arial 24'),).pack(padx=10, pady=10)
    Button(newWindow, text="Check", height=2,
           width=20, command=lambda: testData()).pack(padx=100, pady=20)


def showMessage(stringVal):
    messagebox.showinfo("Warning",  stringVal)


def testData():
    aadharNewValue = NewaadharValue.get()
    values = getValues()
    if not aadharNewValue.isdigit():
        showMessage("Please enter correct aadhar number")

        quit()
    if len(aadharNewValue) != 12:
        showMessage("Please enter 12 digit aadhar number")

    if aadharNewValue in values:
        showMessage("Eligible for the exam")
    else:
        showMessage("Please Register your aadhar no")


window = Tk()
window.config(background="black")
window.geometry('1000x600')

aadharValue = StringVar(window)
NewaadharValue = StringVar(window)


heading_stlye = tkFont.Font(family="Lucida Grande", size=30)
subheading_stlye = tkFont.Font(family="Lucida Grande", size=15)
number_stlye = tkFont.Font(family="Lucida Grande", size=15)
aadharListView = Listbox(window, height=10, width=20, font=number_stlye)
existingValues = getValues()
print(len(existingValues))
if len(existingValues) != 0:
    for x in range(len(existingValues)):
        valueExcel = existingValues[x]
        aadharListView.insert(x, valueExcel)

heading = Label(
    text="Aadhar Registration",
    foreground="white",
    background="black",
    width=50,
    font=heading_stlye
)
subheading = Label(
    text="Please Enter Aadhar No",
    foreground="white",
    background="black",
    width=50,
    height=1,
    font=subheading_stlye
)
aadharNo = Label(
    text="Aadhar List",
    foreground="white",
    background="black",
    width=30,
    font=subheading_stlye
)

enterNumber = Entry(
    window, width=20, textvariable=aadharValue, font=('Arial 24'))

save_button = Button(window, text="Save", height=2, width=20, command=saveData)
next_button = Button(window, text="Next", height=2,
                     width=20, command=nextButtonFun)


if len(existingValues) <= 0:
    heading.pack(padx=50, pady=50)
else:
    heading.pack(padx=50, pady=10)

subheading.pack(padx=100, pady=20)
enterNumber.pack(padx=10, pady=10)
save_button.pack(padx=100, pady=20)

# if sheet.max_row != 1:
if len(existingValues) != 0:
    aadharNo.pack(padx=10, )
    aadharListView.pack(padx=20, pady=10)
next_button.place(x=650, y=510)


window.mainloop()
